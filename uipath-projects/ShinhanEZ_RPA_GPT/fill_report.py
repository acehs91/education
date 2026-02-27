#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fill_report.py - ShinhanEZ 보고서 자동 생성기
- Spring Boot API(localhost:8090)에서 데이터 수집
- 엑셀 템플릿(ShinhanEZ 요약보고서 template.xlsx)에 데이터 삽입
  · 계약 raw data (A8~O?, 15컬럼)
  · 납입 raw data (A8~F?, 6컬럼)
  · 청구 raw data (A8~G?, 7컬럼)
  · GTP요청 데이터  (K6~U?, 11컬럼)
- ShinhanEZ_Report.xlsx 저장 (차트/피벗 자동 반영)
"""

import sys, os, shutil, subprocess
from datetime import datetime, timedelta

# ── 패키지 자동 설치 ─────────────────────────────────────
for _pkg in ("openpyxl", "requests"):
    try:
        __import__(_pkg)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", _pkg, "-q"])

import openpyxl
import requests

# ── 설정 ─────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE   = r"C:\Users\i7C-82\Documents\카카오톡 받은 파일\ShinhanEZ 요약보고서 template.xlsx"
OUTPUT     = os.path.join(SCRIPT_DIR, "ShinhanEZ_Report.xlsx")
API_BASE   = "http://localhost:8090"

# ── 유틸 ─────────────────────────────────────────────────
def fetch(endpoint):
    resp = requests.get(f"{API_BASE}{endpoint}", timeout=30)
    resp.raise_for_status()
    return resp.json().get("data", [])

def parse_date(val):
    if not val:
        return None
    try:
        return datetime.strptime(str(val)[:10], "%Y-%m-%d")
    except Exception:
        return None

def set_date(ws, row, col, dt):
    if dt:
        ws.cell(row, col).value = dt
        ws.cell(row, col).number_format = "YYYY-MM-DD"

# ── 메인 ─────────────────────────────────────────────────
def main():
    # 1) 템플릿 복사
    print(f"[1/7] 템플릿 복사 중...")
    if not os.path.exists(TEMPLATE):
        print(f"  오류: 템플릿을 찾을 수 없습니다 → {TEMPLATE}")
        sys.exit(1)
    shutil.copy2(TEMPLATE, OUTPUT)
    print(f"  → {OUTPUT}")

    # 2) API 수집
    print("[2/7] Spring Boot API 데이터 수집 중...")
    contracts  = fetch("/api/contracts")
    customers  = fetch("/api/customers")
    payments   = fetch("/api/payments")
    claims     = fetch("/api/claims")
    churn_data = requests.get(f"{API_BASE}/api/churn-prediction", timeout=30).json().get("data", [])
    print(f"  계약:{len(contracts)}건  고객:{len(customers)}명  납입:{len(payments)}건  청구:{len(claims)}건  해지분석:{len(churn_data)}명")

    # 3) 인덱스 구성
    cust_map  = {str(c.get("customerId","")): c for c in customers}
    churn_map = {str(c.get("customerId","")): c for c in churn_data}

    pay_by_cid   = {}
    for p in payments:
        pay_by_cid.setdefault(str(p.get("contractId","")), []).append(p)

    claim_by_cid = {}
    for cl in claims:
        claim_by_cid.setdefault(str(cl.get("contractId","")), []).append(cl)

    now               = datetime.now()
    six_months_ago    = now - timedelta(days=180)
    three_months_ago  = now - timedelta(days=90)

    # 4) Excel 로드
    print("[3/7] Excel 파일 로드 중...")
    wb = openpyxl.load_workbook(OUTPUT)
    print(f"  시트: {wb.sheetnames}")

    # ────────────────────────────────────────────────────
    # [4] 계약 raw data  — A7:O? (headers row7, data row8+)
    # ────────────────────────────────────────────────────
    print("[4/7] 계약 raw data 기입 중...")
    ws = wb["계약 raw data"]

    for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    for i, c in enumerate(contracts):
        r    = 8 + i
        cust = cust_map.get(str(c.get("customerId","")), {})

        ws.cell(r, 1).value  = c.get("contractId")
        ws.cell(r, 2).value  = c.get("customerId")
        ws.cell(r, 3).value  = cust.get("name") or c.get("customerName")
        set_date(ws, r, 4, parse_date(cust.get("birthDate")))
        ws.cell(r, 5).value  = cust.get("gender")
        ws.cell(r, 6).value  = c.get("productId")
        ws.cell(r, 7).value  = c.get("productName")
        ws.cell(r, 8).value  = c.get("contractCoverage")
        set_date(ws, r, 9,  parse_date(c.get("regDate")))
        set_date(ws, r, 10, parse_date(c.get("expiredDate")))
        set_date(ws, r, 11, parse_date(c.get("updateDate")))
        ws.cell(r, 12).value = c.get("premiumAmount")
        ws.cell(r, 13).value = c.get("paymentCycle")
        ws.cell(r, 14).value = c.get("contractStatus")
        ws.cell(r, 15).value = c.get("premiumAmount")   # 납입 시 보험료 = 월보험료

    # 테이블 범위 업데이트
    for tbl in ws.tables.values():
        tbl.ref = f"A7:O{max(8, 7 + len(contracts))}"

    # ────────────────────────────────────────────────────
    # [5] 납입 raw data  — A7:F? (headers row7, data row8+)
    # ────────────────────────────────────────────────────
    print("[5/7] 납입 raw data 기입 중...")
    ws = wb["납입 raw data"]

    for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    for i, p in enumerate(payments):
        r = 8 + i
        ws.cell(r, 1).value = p.get("paymentId")
        ws.cell(r, 2).value = p.get("contractId")
        set_date(ws, r, 3, parse_date(p.get("paymentDate")))
        set_date(ws, r, 4, parse_date(p.get("dueDate")))
        ws.cell(r, 5).value = p.get("amount")
        ws.cell(r, 6).value = p.get("status")

    for tbl in ws.tables.values():
        tbl.ref = f"A7:F{max(8, 7 + len(payments))}"

    # ────────────────────────────────────────────────────
    # [6] 청구 raw data  — A7:G? (headers row7, data row8+)
    # ────────────────────────────────────────────────────
    print("[6/7] 청구 raw data 기입 중...")
    ws = wb["청구 raw data"]

    for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    for i, cl in enumerate(claims):
        r = 8 + i
        ws.cell(r, 1).value = cl.get("claimId")
        ws.cell(r, 2).value = cl.get("contractId")
        set_date(ws, r, 3, parse_date(cl.get("claimDate")))
        ws.cell(r, 4).value = cl.get("claimAmount")
        set_date(ws, r, 5, parse_date(cl.get("paidAt")))
        ws.cell(r, 6).value = cl.get("paidAmount")
        ws.cell(r, 7).value = cl.get("status")

    for tbl in ws.tables.values():
        tbl.ref = f"A7:G{max(8, 7 + len(claims))}"

    # ────────────────────────────────────────────────────
    # [6b] GTP요청 데이터 — K5:U? (headers row5, data row6+)
    #   K=col11 계약번호  L=col12 고객이름  M=col13 생년월일
    #   N=col14 계약일    O=col15 보험료    P=col16 3개월연체횟수
    #   Q=col17 총연체건수 R=col18 납입완료비율 S=col19 청구건수
    #   T=col20 6개월청구여부 U=col21 해지확률
    # ────────────────────────────────────────────────────
    print("[6b] GTP요청 데이터 기입 중...")
    ws = wb["GTP요청 데이터"]

    # 기존 데이터 제거 (K-U 컬럼만)
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=11, max_col=21):
        for cell in row:
            cell.value = None

    for i, c in enumerate(contracts):
        r    = 6 + i
        cid  = str(c.get("contractId",""))
        cuid = str(c.get("customerId",""))
        cust = cust_map.get(cuid, {})
        pays = pay_by_cid.get(cid, [])
        cls_ = claim_by_cid.get(cid, [])
        churn = churn_map.get(cuid, {})

        total_pays    = len(pays)
        paid_pays     = sum(1 for p in pays if p.get("status") == "PAID")
        overdue_total = sum(1 for p in pays if p.get("status") == "OVERDUE")
        overdue_3m    = sum(
            1 for p in pays
            if p.get("status") == "OVERDUE"
            and parse_date(p.get("dueDate"))
            and parse_date(p.get("dueDate")) >= three_months_ago
        )
        pay_ratio  = round(paid_pays / total_pays * 100, 1) if total_pays > 0 else 100.0
        claim_cnt  = len(cls_)
        has_6m     = "Y" if any(
            parse_date(cl.get("claimDate")) and
            parse_date(cl.get("claimDate")) >= six_months_ago
            for cl in cls_
        ) else "N"
        risk_score = churn.get("riskScore", 0)
        churn_prob = round(min(risk_score / 100.0, 1.0), 4)

        ws.cell(r, 11).value = c.get("contractId")
        ws.cell(r, 12).value = cust.get("name") or c.get("customerName")
        set_date(ws, r, 13, parse_date(cust.get("birthDate")))
        set_date(ws, r, 14, parse_date(c.get("regDate")))
        ws.cell(r, 15).value = c.get("premiumAmount")
        ws.cell(r, 16).value = overdue_3m
        ws.cell(r, 17).value = overdue_total
        ws.cell(r, 18).value = pay_ratio
        ws.cell(r, 19).value = claim_cnt
        ws.cell(r, 20).value = has_6m
        ws.cell(r, 21).value = churn_prob

    for tbl in ws.tables.values():
        tbl.ref = f"K5:U{max(6, 5 + len(contracts))}"

    # 7) 저장
    print("[7/7] Excel 저장 중...")
    wb.save(OUTPUT)
    print(f"  완료! → {OUTPUT}")
    print(f"  생성일시: {now.strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    try:
        main()
    except requests.exceptions.ConnectionError:
        print("오류: Spring Boot 서버(localhost:8090)에 연결할 수 없습니다.", file=sys.stderr)
        print("서버가 실행 중인지 확인해주세요.", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        import traceback
        print(f"오류 발생: {e}", file=sys.stderr)
        traceback.print_exc()
        sys.exit(1)
